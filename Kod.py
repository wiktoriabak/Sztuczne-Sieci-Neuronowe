import numpy as np
import openpyxl
import itertools
import matplotlib.pyplot as plt
from sklearn.preprocessing import StandardScaler, MinMaxScaler
from sklearn.model_selection import train_test_split


class DeepNeuralNetwork:
    def __init__(self, sizes, activation='relu', optimizer='adam'):
        self.sizes = sizes
        self.optimizer = optimizer

        self.activations = {
            'relu': (self.relu, self.relu_derivative),
            'sigmoid': (self.sigmoid, self.sigmoid_derivative),
            'tanh': (np.tanh, lambda x: 1 - np.tanh(x) ** 2),
            'softplus': (self.softplus, self.softplus_derivative)
        }
        if activation not in self.activations:
            raise ValueError(
                f"Unsupported activation: {activation}. Supported activations: {list(self.activations.keys())}")
        self.activation, self.activation_derivative = self.activations[activation]

        self.params = self.initialize()
        self.cache = {}
        self.opt_cache = self.initialize_optimizer()

    def softplus(self, x):
        return np.log(1 + np.exp(x))

    def softplus_derivative(self, x):
        return 1 / (1 + np.exp(-x))


    def relu(self, x):
        return np.maximum(0, x)

    def relu_derivative(self, x):
        return np.where(x > 0, 1, 0)

    def sigmoid(self, x):
        return 1 / (1 + np.exp(-x))

    def sigmoid_derivative(self, x):
        s = self.sigmoid(x)
        return s * (1 - s)

    def softmax(self, x):
        exps = np.exp(x - np.max(x, axis=0, keepdims=True))
        return exps / np.sum(exps, axis=0, keepdims=True)

    def initialize(self):
        params = {}
        for i in range(1, len(self.sizes)):
            params[f"W{i}"] = np.random.randn(self.sizes[i], self.sizes[i - 1]) * np.sqrt(
                2. / self.sizes[i - 1])
            params[f"b{i}"] = np.zeros((self.sizes[i], 1))
        return params

    def initialize_optimizer(self):
        opt_cache = {}
        if self.optimizer in ['adam', 'nadam']:
            for i in range(1, len(self.sizes)):
                opt_cache[f"vW{i}"] = np.zeros_like(self.params[f"W{i}"])
                opt_cache[f"vb{i}"] = np.zeros_like(self.params[f"b{i}"])
                opt_cache[f"sW{i}"] = np.zeros_like(self.params[f"W{i}"])
                opt_cache[f"sb{i}"] = np.zeros_like(self.params[f"b{i}"])
        elif self.optimizer in ['rmsprop', 'adagrad']:
            for i in range(1, len(self.sizes)):
                opt_cache[f"sW{i}"] = np.zeros_like(self.params[f"W{i}"])
                opt_cache[f"sb{i}"] = np.zeros_like(self.params[f"b{i}"])
        return opt_cache

    def feed_forward(self, x):
        self.cache["A0"] = x.T
        for i in range(1, len(self.sizes)):
            self.cache[f"Z{i}"] = np.dot(self.params[f"W{i}"], self.cache[f"A{i - 1}"]) + self.params[f"b{i}"]
            if i < len(self.sizes) - 1:
                self.cache[f"A{i}"] = self.activation(self.cache[f"Z{i}"])
            else:
                self.cache[f"A{i}"] = self.softmax(self.cache[f"Z{i}"])
        return self.cache[f"A{len(self.sizes) - 1}"]

    def back_propagate(self, y, output):
        grads = {}
        L = len(self.sizes) - 1
        m = y.shape[0]
        dZ = output - y.T
        for i in reversed(range(1, L + 1)):
            grads[f"W{i}"] = (1 / m) * np.dot(dZ, self.cache[f"A{i - 1}"].T)
            grads[f"b{i}"] = (1 / m) * np.sum(dZ, axis=1, keepdims=True)
            if i > 1:
                dZ = np.dot(self.params[f"W{i}"].T, dZ) * self.activation_derivative(self.cache[f"Z{i - 1}"])
        self.grads = grads

    def mean_absolute_error(self, y, output):
        return np.mean(np.abs(y - output.T))

    def mean_squared_error(self, y, output):
        return np.mean((y - output.T) ** 2)

    def optimize(self, l_rate, beta1=0.9, beta2=0.999, epsilon=1e-8):
        if self.optimizer in ['adam', 'nadam']:
            for i in range(1, len(self.sizes)):
                self.opt_cache[f"vW{i}"] = beta1 * self.opt_cache[f"vW{i}"] + (1 - beta1) * self.grads[f"W{i}"]
                self.opt_cache[f"vb{i}"] = beta1 * self.opt_cache[f"vb{i}"] + (1 - beta1) * self.grads[f"b{i}"]
                self.opt_cache[f"sW{i}"] = beta2 * self.opt_cache[f"sW{i}"] + (1 - beta2) * (self.grads[f"W{i}"] ** 2)
                self.opt_cache[f"sb{i}"] = beta2 * self.opt_cache[f"sb{i}"] + (1 - beta2) * (self.grads[f"b{i}"] ** 2)

                vW_corr = self.opt_cache[f"vW{i}"] / (1 - beta1)
                vb_corr = self.opt_cache[f"vb{i}"] / (1 - beta1)
                sW_corr = self.opt_cache[f"sW{i}"] / (1 - beta2)
                sb_corr = self.opt_cache[f"sb{i}"] / (1 - beta2)

                self.params[f"W{i}"] -= l_rate * vW_corr / (np.sqrt(sW_corr) + epsilon)
                self.params[f"b{i}"] -= l_rate * vb_corr / (np.sqrt(sb_corr) + epsilon)

        elif self.optimizer == 'rmsprop':
            for i in range(1, len(self.sizes)):
                self.opt_cache[f"sW{i}"] = 0.9 * self.opt_cache[f"sW{i}"] + 0.1 * (self.grads[f"W{i}"] ** 2)
                self.opt_cache[f"sb{i}"] = 0.9 * self.opt_cache[f"sb{i}"] + 0.1 * (self.grads[f"b{i}"] ** 2)

                self.params[f"W{i}"] -= l_rate * self.grads[f"W{i}"] / (np.sqrt(self.opt_cache[f"sW{i}"]) + epsilon)
                self.params[f"b{i}"] -= l_rate * self.grads[f"b{i}"] / (np.sqrt(self.opt_cache[f"sb{i}"]) + epsilon)

        elif self.optimizer == 'adagrad':
            for i in range(1, len(self.sizes)):
                self.opt_cache[f"sW{i}"] += self.grads[f"W{i}"] ** 2
                self.opt_cache[f"sb{i}"] += self.grads[f"b{i}"] ** 2

                self.params[f"W{i}"] -= l_rate * self.grads[f"W{i}"] / (np.sqrt(self.opt_cache[f"sW{i}"]) + epsilon)
                self.params[f"b{i}"] -= l_rate * self.grads[f"b{i}"] / (np.sqrt(self.opt_cache[f"sb{i}"]) + epsilon)

    def accuracy(self, y, output):
        return np.mean(np.argmax(y, axis=1) == np.argmax(output.T, axis=1))

    def cross_entropy_loss(self, y, output, epsilon=1e-12):
        m = y.shape[0]
        output = np.clip(output.T, epsilon, 1 - epsilon)
        return -np.sum(y * np.log(output)) / m

    def train_epoch(self, x_train, y_train, batch_size, l_rate):
        num_batches = -(-x_train.shape[0] // batch_size)
        perm = np.random.permutation(x_train.shape[0])
        x_train_shuffled = x_train[perm]
        y_train_shuffled = y_train[perm]

        for batch in range(num_batches):
            start = batch * batch_size
            end = min(start + batch_size, x_train.shape[0])
            x_batch = x_train_shuffled[start:end]
            y_batch = y_train_shuffled[start:end]

            output = self.feed_forward(x_batch)
            self.back_propagate(y_batch, output)
            self.optimize(l_rate)

    def train(self, x_train, y_train, x_test, y_test, epochs, batch_size, l_rate):
        epoch_results = []

        for epoch in range(epochs):
            self.train_epoch(x_train, y_train, batch_size, l_rate)

            train_output = self.feed_forward(x_train)
            train_acc = self.accuracy(y_train, train_output)
            train_loss = self.cross_entropy_loss(y_train, train_output)
            train_mse = self.mean_squared_error(y_train, train_output)
            train_mae = self.mean_absolute_error(y_train, train_output)

            test_output = self.feed_forward(x_test)
            test_acc = self.accuracy(y_test, test_output)
            test_loss = self.cross_entropy_loss(y_test, test_output)
            test_mse = self.mean_squared_error(y_test, test_output)
            test_mae = self.mean_absolute_error(y_test, test_output)

            print(f"Epoch {epoch + 1}/{epochs}: "
                  f"Train Accuracy = {train_acc:.4f}, Train Loss = {train_loss:.4f}, Train MSE = {train_mse:.4f}, Train MAE = {train_mae:.4f}, "
                  f"Test Accuracy = {test_acc:.4f}, Test Loss = {test_loss:.4f}, Test MSE = {test_mse:.4f}, Test MAE = {test_mae:.4f}")

            epoch_results.append({
                "epoch": epoch + 1,
                "train_accuracy": train_acc,
                "train_loss": train_loss,
                "train_mse": train_mse,
                "train_mae": train_mae,
                "test_accuracy": test_acc,
                "test_loss": test_loss,
                "test_mse": test_mse,
                "test_mae": test_mae
            })

        return epoch_results


def load_data(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return np.array(data)


def preprocess_data(data, scaling_method='standard'):
    X = data[:, :-1].astype(float)
    y = data[:, -1].astype(int)

    if scaling_method == 'standard':
        scaler = StandardScaler()
    elif scaling_method == 'minmax':
        scaler = MinMaxScaler()
    else:
        raise ValueError(f"Unsupported scaling method: {scaling_method}")

    X_scaled = scaler.fit_transform(X)

    y_one_hot = np.zeros((y.size, y.max() + 1))
    y_one_hot[np.arange(y.size), y] = 1
    return X_scaled, y_one_hot


def plot_training_history(epoch_results):
    # Wykres dokładności
    plt.figure(figsize=(10, 6))
    train_acc = [result['train_accuracy'] for result in epoch_results]
    test_acc = [result['test_accuracy'] for result in epoch_results]

    plt.plot(range(1, len(epoch_results) + 1), train_acc, label='Train Accuracy')
    plt.plot(range(1, len(epoch_results) + 1), test_acc, label='Test Accuracy')
    plt.title('Model Accuracy')
    plt.ylabel('Accuracy')
    plt.xlabel('Epoch')
    plt.legend(loc='upper left')
    plt.show()

    # Wykres straty
    plt.figure(figsize=(10, 6))
    train_loss = [result['train_loss'] for result in epoch_results]
    test_loss = [result['test_loss'] for result in epoch_results]

    plt.plot(range(1, len(epoch_results) + 1), train_loss, label='Train Loss')
    plt.plot(range(1, len(epoch_results) + 1), test_loss, label='Test Loss')
    plt.title('Model Loss')
    plt.ylabel('Loss')
    plt.xlabel('Epoch')
    plt.legend(loc='upper left')
    plt.show()

    # Wykres MSE
    plt.figure(figsize=(10, 6))
    train_mse = [result['train_mse'] for result in epoch_results]
    test_mse = [result['test_mse'] for result in epoch_results]

    plt.plot(range(1, len(epoch_results) + 1), train_mse, label='Train MSE')
    plt.plot(range(1, len(epoch_results) + 1), test_mse, label='Test MSE')
    plt.title('Model Mean Squared Error (MSE)')
    plt.ylabel('MSE')
    plt.xlabel('Epoch')
    plt.legend(loc='upper left')
    plt.show()

    # Wykres MAE
    plt.figure(figsize=(10, 6))
    train_mae = [result['train_mae'] for result in epoch_results]
    test_mae = [result['test_mae'] for result in epoch_results]

    plt.plot(range(1, len(epoch_results) + 1), train_mae, label='Train MAE')
    plt.plot(range(1, len(epoch_results) + 1), test_mae, label='Test MAE')
    plt.title('Model Mean Absolute Error (MAE)')
    plt.ylabel('MAE')
    plt.xlabel('Epoch')
    plt.legend(loc='upper left')
    plt.show()


def parameter_search(filepath, param_grid):
    data = load_data(filepath)
    X, y = preprocess_data(data, scaling_method='standard')

    results = []
    epoch_results = []

    param_combinations = [dict(zip(param_grid.keys(), values)) for values in itertools.product(*param_grid.values())]

    for params in param_combinations:
        print(f"Testing parameters: {params}")

        x_train, x_test, y_train, y_test = train_test_split(X, y, test_size=params["test_size"], random_state=42)
        sizes = [X.shape[1]] + [params['neurons']] * params['layers'] + [y.shape[1]]

        model = DeepNeuralNetwork(sizes, activation=params['activation'], optimizer=params['learning_method'])

        epoch_result = model.train(x_train, y_train, x_test, y_test, epochs=100, batch_size=32, l_rate=0.01)

        results.append({
            **params,
            'train_accuracy': epoch_result[-1]['train_accuracy'],
            'train_loss': epoch_result[-1]['train_loss'],
            'train_mse': epoch_result[-1]['train_mse'],
            'train_mae': epoch_result[-1]['train_mae'],
            'test_accuracy': epoch_result[-1]['test_accuracy'],
            'test_loss': epoch_result[-1]['test_loss'],
            'test_mse': epoch_result[-1]['test_mse'],
            'test_mae': epoch_result[-1]['test_mae']
        })

        for epoch in epoch_result:
            epoch_results.append({
                **params,
                "epoch": epoch["epoch"],
                "train_accuracy": epoch["train_accuracy"],
                "train_loss": epoch["train_loss"],
                "test_accuracy": epoch["test_accuracy"],
                "test_loss": epoch["test_loss"],
                "train_mse": epoch["train_mse"],
                "test_mse": epoch["test_mse"],
                "train_mae": epoch["train_mae"],
                "test_mae": epoch["test_mae"]
            })
        plot_training_history(epoch_result)

    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet.title = "Parameter Search Results"

    headers = list(results[0].keys())
    sheet.append(headers)

    for result in results:
        sheet.append([result[key] for key in headers])

    sheet_epochs = workbook.create_sheet(title="Epoch-wise Results")
    epoch_headers = list(epoch_results[0].keys())
    sheet_epochs.append(epoch_headers)

    for epoch_result in epoch_results:
        sheet_epochs.append([epoch_result[key] for key in epoch_headers])

    workbook.save("C:/Users/laura_a1y0snp/Desktop/5 semestr/IO/sieci/results.xlsx")
    print("Results saved to 'results.xlsx'")

param_grid = {
    "neurons": [8], # 4, 16, 32
    "layers": [2], # 1, 3, 4
    "learning_method": ["adam"],  # , "rmsprop", "adagrad", "nadam"
    "activation": ["relu"], #  sigmoid, tanh, softplus
    "test_size": [0.3] # 0.1, 0.2, 0.4
}

parameter_search("C:/Users/laura_a1y0snp/Desktop/5 semestr/IO/sieci/diabietes_dane.xlsx", param_grid)